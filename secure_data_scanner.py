"""
AI Sensitive Data Scanner with Masked Output
Masks sensitive values in the final Excel report.
"""

import os
import re
from textblob import TextBlob
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# ------------------------------
# CONFIGURATION
# ------------------------------
SCAN_FOLDER = r"C:\AI-Training\ai-programmers\week_7"
REPORT_FILE = os.path.join(SCAN_FOLDER, "Sensitive_Data_AI_Report.xlsx")

SENSITIVE_PATTERNS = {
    "Email Address": r"\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}\b",
    "Credit Card": r"\b(?:\d[ -]*?){13,16}\b",
    "SSN": r"\b\d{3}-\d{2}-\d{4}\b",
    "API Key": r"(?i)(api[_-]?key|token|secret)\s*[:=]\s*['\"]?[A-Za-z0-9\-_]{8,}['\"]?",
    "Password": r"(?i)(password|passwd|pwd)\s*[:=]\s*['\"].+?['\"]"
}

SUPPORTED_EXTENSIONS = {
    '.txt', '.csv', '.log', '.json', '.xml', '.yaml', '.yml',
    '.ini', '.config', '.py', '.js', '.java', '.env'
}


def mask_sensitive_data(text, data_type):
    """Mask sensitive values while preserving limited context."""
    if data_type == "Email Address":
        parts = text.split('@')
        if len(parts) == 2:
            username = parts[0]
            domain = parts[1]
            masked_user = username[:2] + '*' * max(len(username) - 2, 0)
            return f"{masked_user}@{domain}"

    elif data_type == "Credit Card":
        digits = re.sub(r'\D', '', text)
        if len(digits) >= 4:
            return '*' * (len(digits) - 4) + digits[-4:]

    elif data_type == "SSN":
        return '***-**-' + text[-4:]

    elif data_type in {"API Key", "Password"}:
        return text[:4] + '*' * max(len(text) - 4, 0)

    return '***MASKED***'


def analyze_risk(text):
    sentiment = TextBlob(text).sentiment.polarity
    if sentiment < -0.2:
        return "High"
    elif sentiment < 0.2:
        return "Medium"
    return "Low"


def scan_file(file_path):
    findings = []

    try:
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as file:
            for line_num, line in enumerate(file, start=1):
                for data_type, pattern in SENSITIVE_PATTERNS.items():
                    for match in re.finditer(pattern, line):
                        original_text = match.group(0)
                        findings.append({
                            'File Name': os.path.basename(file_path),
                            'File Path': file_path,
                            'Line Number': line_num,
                            'Data Type': data_type,
                            'Matched Text': mask_sensitive_data(original_text, data_type),
                            'Risk Level': analyze_risk(line)
                        })
    except Exception as e:
        print(f"Skipped {file_path}: {e}")

    return findings


def scan_folder(folder_path):
    all_findings = []
    for root, _, files in os.walk(folder_path):
        for file in files:
            if os.path.splitext(file)[1].lower() in SUPPORTED_EXTENSIONS:
                all_findings.extend(scan_file(os.path.join(root, file)))
    return all_findings


def export_to_excel(findings, output_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sensitive Data Report"

    headers = [
        'File Name', 'File Path', 'Line Number',
        'Data Type', 'Matched Text', 'Risk Level'
    ]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='1F4E78')

    for item in findings:
        ws.append([
            item['File Name'],
            item['File Path'],
            item['Line Number'],
            item['Data Type'],
            item['Matched Text'],
            item['Risk Level']
        ])

    for column in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in column)
        ws.column_dimensions[column[0].column_letter].width = max_length + 2

    wb.save(output_file)


def main():
    print("AI Sensitive Data Scanner")
    print("Scanning files...\n")

    findings = scan_folder(SCAN_FOLDER)

    if findings:
        export_to_excel(findings, REPORT_FILE)
        print(f"Scan complete. Report saved to: {REPORT_FILE}")
        print(f"Total findings: {len(findings)}")
    else:
        print("No sensitive data found.")


if __name__ == '__main__':
    main()
